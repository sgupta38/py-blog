# Spreadsheet handling via python
# Columns[letters] & Row[numbers]
## todo: Writing to Spreadsheet

import openpyxl
import os

#open workbook
workbook = openpyxl.load_workbook('Shapeshifter.xlsx') ## Take ur excelsheet and change accordingly
print(type(workbook))

#open sheet
sheet = workbook.get_sheet_by_name('Gen4-Sprint 3.3')
print(type(sheet))
print(workbook.get_sheet_names())
print(sheet['c33'])  # This returns the cell object
print(sheet['c33'].value)  # This returns the cell value

print(sheet['G33'].value)  # This returns the cell value [integer format]
print(str(sheet['G33'].value))  # This returns the cell value [string format]

print(sheet.cell(row=4, column=3)) # Note that row and columns begins from 1: This returns the object
print(sheet['h4'].value)

## Note: Its better to use "sheet.cell(row=x, column=y).value" format whenever we want to retrieve bunch of data
##
##for example:

for i in range(52, 63):
    print(i, sheet.cell(row=i, column=3).value, sheet.cell(row=i, column=4).value)

print()
print()
##########################################################################################################
##                          Creating and modifying Spreadsheet
##########################################################################################################

wb = openpyxl.Workbook() # This creates new workbook
print(wb.get_sheet_names())

sheet = wb.get_sheet_by_name('Sheet')
sheet['a1'] = 46;
sheet['a2'] = 'Sonu Gupta';

## Till now whatever we have done is saved in 'Memory'.
# To save it on hard drive use 'save()'

wb.save('mysheet1.xlsx')

# Lets open the workbook for modifying
wb = openpyxl.load_workbook('mysheet1.xlsx')

#To add new Sheet
sheet2 = wb.create_sheet()          # by default added to last
print(wb.get_sheet_names())

## If we want to change the name of sheet

print(sheet2.title)
sheet2.title = 'My New sheet name'
print(wb.get_sheet_names())

wb.save('mysheet1.xlsx')

## To add at particular location
sheet2 = wb.create_sheet(index=0, title='My Other sheet')          # by default added to last
wb.save('mysheet1.xlsx')
